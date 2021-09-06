import sys

import openpyxl as xl
from openpyxl.chart import Reference,  LineChart
# import seaborn as sns; sns.set()
import numpy as np
from sklearn.cluster import KMeans
import pandas as pd
import os

# Отображение колонок при выводе данных
pd.set_option('display.max_columns', 7)


class Cluster():

    def __init__(self, df_initial, n_cluster):
        self.df_initial=df_initial
        self.n_cluster=n_cluster
        self.df_new = None
        self.df_cluster_for_plot = None
        self.df_agg = None

    def create_line_chart_excel(self, name_x_axis, name_y_axis, col_x, col_y, file_name):
        workbook = xl.load_workbook(f'Результаты расчета/{file_name}')
        sheet_1 = workbook['Кластеризация']
        values = Reference(sheet_1, min_row=1, max_row=sheet_1.max_row, min_col=col_y, max_col=col_y)
        index = Reference(sheet_1, min_row=2, max_row=sheet_1.max_row, min_col=col_x)
        chart = LineChart()
        chart.y_axis.title = name_y_axis
        chart.x_axis.title = name_x_axis
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(index)
        sheet_1.add_chart(chart, 'I2')
        workbook.save(f'Результаты расчета/{file_name}')

    def _change(self, df, name):
        df['RollT'] = np.roll(df, 1)
        df['Шаг времени, ч'] = pd.to_timedelta(df.iloc[1, 0] - df.iloc[1, 1]).total_seconds() / 3600
        df['Минута'] = df[name].apply(lambda x: x.minute)
        df['Час'] = df.apply(lambda x: round(x[name].hour + x['Минута'] / 60, 1), axis=1)
        df.drop(columns = ['RollT', 'Минута'], inplace=True)

        return df

    def fun(self):
        df = self.df_initial.dropna()
        print(df.dtypes)
        name_list = df.columns
        if len(name_list) > 2:
            return print('Количество переменных должно быть не более двух.')
        elif df.iloc[:, 0].dtypes == 'datetime64[ns]' or df.iloc[:, 1].dtypes == 'datetime64[ns]':
            name_x = df.columns[0]
            name_y = df.columns[1]
            X = pd.DataFrame().to_numpy()
            if df[name_x].dtypes == 'datetime64[ns]':
                df_x = self._change(df=df[[name_x]], name=name_x)
                df_y = df[[name_y]]
                df = pd.concat([df_x, df_y], axis=1)
                X = df.iloc[:, [2, 3]].to_numpy()
            elif df[name_y].dtypes == 'datetime64[ns]':
                df_y = self._change(df=df[[name_y]], name=name_y)
                df_x = df[[name_x]]
                df = pd.concat([df_y, df_x], axis=1)
                X = df.iloc[:, [3, 2]].to_numpy()

            kmeans = KMeans(n_clusters=self.n_cluster)
            kmeans.fit_predict(X)
            #Класстер сместили на 1 для удосбтва восприятия. Начинается не с нуля
            y_kmeans = np.array(list(map(lambda x: x+1, kmeans.predict(X))))
            # y_kmeans = kmeans.predict(X)

            predict = pd.DataFrame(y_kmeans, columns=['Номер кластера'])
            df_result = pd.concat([df, predict], axis=1)
            df_result = df_result.set_index('Номер кластера')
            centers = kmeans.cluster_centers_
            df_center = pd.DataFrame(centers, columns=[f'Центроида: {name_x}',
                                                       f'Центроида: {name_y}'])
            clusters_list = [i+1 for i in range(self.n_cluster)]
            df_clust = pd.DataFrame(clusters_list, columns=['Номер кластера'])
            df_center = pd.concat([round(df_center, 1), df_clust], axis=1).set_index('Номер кластера')

            df_new = df_result.join(df_center).reset_index()
            df_new.sort_values(by=['Час'], inplace=True)
            df_new['Порядковый номер'] = np.arange(len(df_new))


            self.df_new = df_new.iloc[:, [1, 2, 3, 0, 6, 4, 5]]
            self.df_cluster_for_plot = self.df_new.iloc[:, [2, 4, 2, 5]]
            print(self.df_cluster_for_plot.columns[0])


            if df['Шаг времени, ч'].mean() >= 24:

                pass





                X = df.iloc[:, [1, 4]].to_numpy()

                kmeans = KMeans(n_clusters=self.n_cluster)
                kmeans.fit_predict(X)
                y_kmeans = kmeans.predict(X)

                predict = pd.DataFrame(y_kmeans, columns=['Номер кластера'])

                df_result = pd.concat([df, predict], axis=1)
                df_result = df_result.drop(columns=['dT, ч', 'Минута'])
                df_result = df_result.set_index('Номер кластера')
                centers = kmeans.cluster_centers_
                df_center = pd.DataFrame(centers, columns=[f'Центроида: {name_list[1]}',
                                                           f'Центроида: {name_list[0]}'])
                clusters_list = [i for i in range(self.n_cluster)]
                df_clust = pd.DataFrame(clusters_list, columns=['Номер кластера'])
                df_center = pd.concat([round(df_center, 1), df_clust], axis=1).set_index('Номер кластера')
                df_new = df_result.join(df_center).reset_index()
                df_new.sort_values(by=[f'Центроида: {name_list[1]}'], inplace=True)
                df_new['Порядковый номер'] = np.arange(len(df_new))
                self.df_new = df_new.iloc[:, [1, 2, 3, 0, 6, 4, 5]]
                self.df_cluster_for_plot = self.df_new.iloc[:, [4, 1, 4, 5]]

                # Групповые данные
                df_agg = df_new.groupby(['Номер кластера']).aggregate(
                    {'Порядковый номер': 'count', f'Центроида: {name_list[1]}': 'mean',
                     f'Центроида: {name_list[0]}': 'mean'}).reset_index()
                self.df_agg = df_agg.rename(columns={'Порядковый номер': 'Количество значений попавших в диапазон'})

        elif len(df.columns) == 1:

            df.sort_values(by=[df.iloc[:, [0]].columns[0]], inplace=True, ascending=False)
            df['Уникальный номер'] = np.arange(len(df))

            X = df.iloc[:, [0, 1]].to_numpy()
            kmeans = KMeans(n_clusters=self.n_cluster)
            kmeans.fit_predict(X)
            y_kmeans = kmeans.predict(X)
            predict = pd.DataFrame(y_kmeans, columns=['Номер кластера'])

            df_result = pd.concat([df, predict], axis=1)
            df_result = df_result.set_index('Номер кластера')
            centers = kmeans.cluster_centers_
            df_center = pd.DataFrame(centers, columns=[f'Центроида: {df.iloc[:, [0]].columns[0]}',
                                                       f'Центроида: {df.iloc[:, [1]].columns[0]}'])
            clusters_list = [i for i in range(self.n_cluster)]
            df_clust = pd.DataFrame(clusters_list, columns=['Номер кластера'])
            df_center = pd.concat([round(df_center, 1), df_clust], axis=1).set_index('Номер кластера')
            df_new = df_result.join(df_center).reset_index()
            df_new.sort_values(by=[f'Центроида: {df.iloc[:, [0]].columns[0]}'], inplace=True,  ascending=False )
            df_new['Порядковый номер'] = np.arange(len(df_new))
            self.df_new = df_new.iloc[:, [0, 1, 5, 3]]
            self.df_cluster_for_plot = self.df_new.iloc[:, [2, 1, 2, 3]]

            # # Групповые данные
            df_agg = df_new.groupby(['Номер кластера']).aggregate(
                {'Порядковый номер': 'count', f'Центроида: {df.iloc[:, [0]].columns[0]}': 'mean',
                 }).reset_index()
            self.df_agg = df_agg.rename(columns={'Порядковый номер': 'Количество значений попавших в диапазон'})

        else:
            try:
                X = df.iloc[:, [0, 1]].to_numpy()
                kmeans = KMeans(n_clusters=self.n_cluster)
                kmeans.fit_predict(X)
                y_kmeans = kmeans.predict(X)
                predict = pd.DataFrame(y_kmeans, columns=['Номер кластера'])

                df_result = pd.concat([df, predict], axis=1)
                df_result = df_result.set_index('Номер кластера')
                centers = kmeans.cluster_centers_
                df_center = pd.DataFrame(centers, columns=[f'Центроида: {name_list[0]}',
                                                           f'Центроида: {name_list[1]}'])
                clusters_list = [i for i in range(self.n_cluster)]
                df_clust = pd.DataFrame(clusters_list, columns=['Номер кластера'])
                df_center = pd.concat([round(df_center, 1), df_clust], axis=1).set_index('Номер кластера')
                df_new = df_result.join(df_center).reset_index()
                df_new.sort_values(by=[f'Центроида: {name_list[1]}'], inplace=True, ascending=False)
                df_new['Порядковый номер'] = np.arange(len(df_new))


                self.df_new = df_new.iloc[:, [1, 2, 0, 5, 3, 4]]
                self.df_cluster_for_plot = self.df_new.iloc[:, [1, 0, 5, 4]]
                # # Групповые данные
                df_agg = df_new.groupby(['Номер кластера']).aggregate(
                    {'Порядковый номер': 'count', f'Центроида: {name_list[1]}': 'mean',
                     f'Центроида: {name_list[0]}': 'mean'}).reset_index()
                self.df_agg = df_agg.rename(columns={'Порядковый номер': 'Количество значений попавших в диапазон'})
            except Exception:
               print('Что-то пошло не так.')


path = 'd:\РАБОТА, НАУКА И ПРОЧЕЕ ВАЖНОЕ!\++РАБОТА\++5 ПРОГРАММИРОВАНИЕ\Phaton\Мои Script на Python\Кластеризация streamlit\всякое\Анализируемая статистика.xlsx'

df = pd.read_excel(path)


f = Cluster(df_initial=df, n_cluster=5)
f.fun()
df_new = f.df_cluster_for_plot

df_new.to_excel('gh.xlsx')









